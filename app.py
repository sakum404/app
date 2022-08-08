import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
import openpyxl as exl
from datetime import datetime
from openpyxl.styles import Font ,Alignment,Border,Side
import numpy
date_string = datetime.now()

wb = exl.load_workbook(filename="bsg.xlsx")
ws = wb.active
sheet = wb["BSG-PR"]

wb2 = exl.load_workbook(filename='BSG purchasing index PRO-010-02.xlsx')
ws2 = wb2.active
sheet = wb2["Procurement index"]
nextrow = sheet.max_row + 1
font_style = Font(name='Calibri', sz='10')
alig_style_left = Alignment(wrapText=True, horizontal='left', vertical='center')
alig_style_right = Alignment(wrapText=True, horizontal='right', vertical='center')
alig_style_center = Alignment(wrapText=True, horizontal='center', vertical='center')
top = Side(border_style='thin')
bottom = Side(border_style='thin')
left = Side(border_style='thin')
right = Side(border_style='thin')

border = Border(top=top, left=left, right=right, bottom=bottom)


a1 = sheet.max_row
a2 = ws2[f'B{a1}']
a3 = a2.value

a1 = sheet.max_row
a2 = ws2[f'A{a1}']
a3 = a2.value
string = numpy.array(a3[8:12])
int = string.astype(numpy.int_)
purchase = int + 1



root = tk.Tk()
root.title('Severn purchasing  V 1.0')
root.geometry('600x575')
root.resizable(width=False, height=False)
root.iconbitmap('severn-logo.ico')


#Func for Buttons
def save_as():
    name = combo_name.get()
    purpose = enter_purpose.get()
    name_service = enter_name_service.get(1.0, tk.END+"-1c")
    quan = enter_quan.get()
    person = enter_person.get()
    vendor = enter_vendor.get()
    price = enter_price.get()
    contact = enter_contact.get(1.0, tk.END+"-1c")
    currency = combo_currency.get()
    types = combo_type.get()
    costs1 = combo_cost.get()

        
    if types == 'General':
        type1 = 'G'
    elif types == 'Valve Contract':
        type1 = 'V'
    elif types == 'Pump Contract':
        type1 = 'P'
    elif types == 'Process diagnostics':
        type1 = 'PD'
    elif types == 'Traded products':
        type1 = 'TP'
    elif types == 'Operations Department':
        type1 = 'O'
    elif types == 'City shop':
        type1 = 'C'
    elif types == 'Karabatan shop':
        type1 = 'K'
    elif types == 'D-Island shop':
        type1 = 'D'

    if costs1 == 'Workshop equipment | Оборудование для цеха':
        center = '001'
    elif costs1 == 'Workshop tools | Инструменты для цеха':
        center = '002'
    elif costs1 == 'Workshop consumables | Расходные материалы для цеха':
        center = '003'
    elif costs1 == 'PPE, Safety equipment |СИЗ и оборудование по ТБ':
        center = '004'
    elif costs1 == 'Office equipment and furniture | Офисное оборудование и мебель':
        center = '005'
    elif costs1 == 'Office consumables and stationary | Расходные материалы и канцелярские товары':
        center = 'O06'
    elif costs1 == 'Purchase of services for workshop | Услуги для цеха':
        center = '007'
    elif costs1 == 'Car maintenance and car consumables | Обслуживание автомобилей и расходные материалы к автомобилям':
        center = '008'
    elif costs1 == 'Office maintenance (household expenses, tea, cofffee) | Обеспечение офиса':
        center = '009'
    elif costs1 == 'Repair of fixed assets | Ремонт основных средств':
        center = '010'
    elif costs1 == 'Third-party services | Услуги третьей стороны':
        center = '011'
    elif costs1 == 'Trainings | Тренинги':
        center = '012'
    elif costs1 == 'IT software applications, licenses, etc':
        center = '013'



    ws['E4'] = f'BSG-PR-0{purchase}'
    ws['I4'] = name
    ws['I5'] = purpose
    ws['C19'] = name_service
    ws['E19'] = quan
    ws['F19'] = person
    ws['I19'] = vendor
    ws['M19'] = price
    ws['E5'] = date_string.strftime('%d.%m.%Y')
    ws['L19'] = type1
    try:
        des = f'{center}-{type1}'
        ws['L19'] = des

    except:
        pass



    str1 = numpy.array(quan)
    int1 = str1.astype(numpy.int_)

    str2 = numpy.array(price)
    int2 = str2.astype(numpy.int_)

    TPrice = int1 * int2

    ws2[f'A{nextrow}'].border = border
    ws2[f'A{nextrow}'].value = f'BSG-PR-0{purchase}'
    ws2[f'A{nextrow}'].alignment = alig_style_center
    ws2[f'A{nextrow}'].font = Font(name='Calibri', sz='10', bold=True)

    ws2[f'B{nextrow}'].border = border
    ws2[f'B{nextrow}'].value = date_string.strftime('%d.%m.%Y')
    ws2[f'B{nextrow}'].alignment = alig_style_center
    ws2[f'B{nextrow}'].font = font_style

    ws2[f'C{nextrow}'].border = border
    ws2[f'C{nextrow}'].value = name
    ws2[f'C{nextrow}'].alignment = alig_style_left
    ws2[f'C{nextrow}'].font = font_style

    ws2[f'D{nextrow}'].border = border
    ws2[f'D{nextrow}'].value = name_service
    ws2[f'D{nextrow}'].alignment = alig_style_left
    ws2[f'D{nextrow}'].font = font_style

    ws2[f'E{nextrow}'].border = border
    ws2[f'E{nextrow}'].value = 'For BSG'
    ws2[f'E{nextrow}'].alignment = alig_style_left
    ws2[f'E{nextrow}'].font = font_style

    ws2[f'F{nextrow}'].border = border
    ws2[f'F{nextrow}'].value = vendor
    ws2[f'F{nextrow}'].alignment = alig_style_left
    ws2[f'F{nextrow}'].font = font_style

    ws2[f'G{nextrow}'].border = border
    ws2[f'G{nextrow}'].value = contact
    ws2[f'G{nextrow}'].alignment = alig_style_left
    ws2[f'G{nextrow}'].font = font_style

    ws2[f'H{nextrow}'].border = border
    ws2[f'H{nextrow}'].value = f'KZT {TPrice}'
    ws2[f'H{nextrow}'].alignment = alig_style_right
    ws2[f'H{nextrow}'].font = font_style

    ws2[f'I{nextrow}'].border = border
    ws2[f'I{nextrow}'].value = currency
    ws2[f'I{nextrow}'].alignment = alig_style_right
    ws2[f'I{nextrow}'].font = font_style

    ws2[f'J{nextrow}'].border = border
    ws2[f'J{nextrow}'].value = 'DDP Atyrau'
    ws2[f'J{nextrow}'].alignment = alig_style_center
    ws2[f'J{nextrow}'].font = font_style

    ws2[f'K{nextrow}'].border = border
    ws2[f'K{nextrow}'].value = 'Supply in process'
    ws2[f'K{nextrow}'].alignment = alig_style_center
    ws2[f'K{nextrow}'].font = font_style

    ws2[f'N{nextrow}'].border = border
    ws2[f'N{nextrow}'].value = name
    ws2[f'N{nextrow}'].alignment = alig_style_left
    ws2[f'N{nextrow}'].font = font_style

    wb2.save('BSG purchasing index PRO-010-02.xlsx')
    # wb.save(f'BSG-PR-0{purchase}.xlsx')
    def save_file():
        down_dict = fd.asksaveasfilename(
            filetypes=[('excel', '*.txt')],
            defaultextension='.xlsx',
            initialfile=f'BSG-PR-0{purchase}.xlsx'

            )
        # path = down_dict +f'BSG-PR-0{purchase}.xlsx'
        wb.save(down_dict)
    save_file()


def cancle():
    root.destroy()


cost_center = ['Workshop equipment | Оборудование для цеха',
         'Workshop tools | Инструменты для цеха',
         'Workshop consumables | Расходные материалы для цеха',
         'PPE, Safety equipment |СИЗ и оборудование по ТБ',
         'Office equipment and furniture | Офисное оборудование и мебель',
         'Office consumables and stationary | Расходные материалы и канцелярские товары ',
         'Purchase of services for workshop | Услуги для цеха',
         'Car maintenance and car consumables | Обслуживание автомобилей и расходные материалы к автомобилям',
         'Office maintenance (household expenses, tea, cofffee) | Обеспечение офиса',
         'Repair of fixed assets | Ремонт основных средств',
         'Third-party services | Услуги третьей стороны',
         'Trainings | Тренинги',
         'IT software applications, licenses, etc',
         ]

names = ['Aizada Abtay',
         'Anuarbek Muhammed',
         'Akhmediyar Salykov',
         'Adrian Owen',
         'Aizada Muratova',
         'Andrew Davidson',
         'Berik Gibbatulin',
         'Dias Kuraishov',
         'Lyailya Sarbayeva',
         'Nurlan Zhunussov',
         'John Carter',
         'Akmaral Izteleuova',
         'Adilet Kumarov',
         'Nurgul Mukanova',
         'Salamat Maukenov',
         'Tarasbaev Azamat',
         'Sergey Mazur',
         'Karlygash Lepessova',
         'Nurgul Zhubanova',
         'Aimira Dzhumagalieva',
         'Shynar Ramazanova']

currency = ['KZT',
            'USD',
            'GBR',
            'RUB',
            'EUR']

types = ['General',
         'Valve Contract',
         'Pump Contract',
         'Process diagnostics',
         'Traded products',
         'Operations Department',
         'City shop',
         'Karabatan shop',
         'D-Island shop']

# widgets
label_name = tk.Label(text = 'Name | Имя:')
combo_name = ttk.Combobox(root,width=50, values=names)
label_name.grid(row=0, column=0, pady=10, padx=5)
combo_name.grid(row=0, column=1,columnspan=2, pady=10, padx=5)
combo_name.bind("<<ComboboxSelected>>", save_as)
combo_name.current(0)

label_purpose = tk.Label(text = 'Purpose | Цель:')
enter_purpose= tk.Entry(root, width=50,font="Arial 10")
label_purpose.grid(row=1, column=0, pady=10, padx=5)
enter_purpose.grid(row=1, column=1,columnspan=2, pady=10, padx=5)

label_name_service = tk.Label(text='Description of materials | services \nОписание материалов | услуг')
enter_name_service = tk.Text(root, width=50, height=3,font="Arial 10")
label_name_service.grid(row=2, column=0, pady=10, padx=5)
enter_name_service.grid(row=2, column=1,columnspan=2, pady=10, padx=5)

label_quan = tk.Label(text='Quantity  | Количество')
enter_quan = tk.Entry(root, width=50, font="Arial 10")
label_quan.grid(row=3, column=0, pady=10, padx=5)
enter_quan.grid(row=3, column=1,columnspan=2, pady=10, padx=5)

label_person = tk.Label(text='Person responsible for materials | \nМатериально ответственное лицо')
enter_person = tk.Entry(root, width=50, font="Arial 10")
label_person.grid(row=4, column=0, pady=10, padx=5)
enter_person.grid(row=4, column=1,columnspan=2, pady=10, padx=5)

label_vendor = tk.Label(text='Recommended vendor | \n  Рекомендуемый поставщик')
enter_vendor = tk.Entry(root, width=50, font="Arial 10")
label_vendor.grid(row=5, column=0, pady=10, padx=5)
enter_vendor.grid(row=5, column=1,columnspan=2, pady=10, padx=5)

label_contact = tk.Label(text = 'Contract / PO / Invoice:')
enter_contact = tk.Text(root, width=50, height=3,font="Arial 10")
label_contact.grid(row=8, column=0, pady=10, padx=5)
enter_contact.grid(row=8, column=1,columnspan=2, pady=10, padx=5)

label_price = tk.Label(text='Unit price | Стоимость за ед.')
enter_price = tk.Entry(root, width=30, font="Arial 10")
combo_currency = ttk.Combobox(root,width=10, values=currency)
label_price.grid(row=6, column=0, pady=10, padx=5)
enter_price.grid(row=6, column=1, rowspan=2, sticky='w',pady=10, padx=5)
combo_currency.grid(row=6, column=2, rowspan=2,sticky='w' )
combo_currency.current(0)

label_type = tk.Label(text='Type')
combo_type = ttk.Combobox(root,width=50, values=types)
label_type.grid(row=9, column=0, pady=10, padx=5)
combo_type.grid(row=9, column=1, columnspan=2, pady=10, padx=5)
combo_type.current(0)

label_cost = tk.Label(text='Cost center | Центр затрат')
combo_cost = ttk.Combobox(root,width=50, values=cost_center)
label_cost.grid(row=10, column=0, pady=10, padx=5)
combo_cost.grid(row=10, column=1, columnspan=2)



btn_save = tk.Button(text='Save as', padx=50, command=save_as)
btn_cancel = tk.Button(text='Cancle',padx=50, command=cancle)
btn_save.grid(row=11, column=1, pady=10 ,columnspan=2, sticky='w')
btn_cancel.grid(row=11, column=2, pady=10,columnspan=2, sticky='w')




root.mainloop()